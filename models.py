from django.utils.translation import ugettext_lazy as _
from django.db import models
from django.urls import reverse
from django.db.models import Min, Max, Sum
from django.utils.functional import SimpleLazyObject
import decimal
from datetime import datetime
from solo.models import SingletonModel
from .tools import valid_iban

class Seizoen(models.Model):
    startjaar = models.PositiveIntegerField(unique=True)
    current = models.BooleanField(default=False)

    @ property
    def eindjaar(self):
        return self.startjaar + 1

    @ property
    def description(self):
        return '%d-%d' % (self.startjaar, self.eindjaar)

    class Meta:
        verbose_name = _("season")
        verbose_name_plural = _("seasons")
        ordering = ['startjaar']

    def __str__(self):
        return self.description

    def save(self, *args, **kwargs):
        "Ensure that only one seizoen is current"
        if self.current:
            # select all other current items
            qs = type(self).objects.filter(current=True)
            # except self (if self already exists)
            if self.pk:
                qs = qs.exclude(pk=self.pk)
            # and deactive them
            qs.update(current=False)

        super(Seizoen, self).save(*args, **kwargs)

class Configuration(SingletonModel):
# when adding fields to this model the --skip-checks option must be added 
# in both the makemigrations and the migrate commands
# e.g.
#
# python manage.py makemigrations --skip-checks
# python manage.py migrate --skip-checks

    last_memberimport = models.DateField(_("last member import"), null=True)
    refresh_all = models.BooleanField(_("refresh all"), default=False)
    seizoen = models.ForeignKey(Seizoen, on_delete=models.PROTECT, null=True)
    mail_username = models.EmailField(blank=True, null=True)
    mail_password = models.CharField(max_length=25, null=True)
    mail_host = models.CharField(max_length=25, null=True)
    mail_port = models.IntegerField(null=True)
    mail_from = models.CharField(max_length=25, null=True)
    mail_use_tls = models.BooleanField(default=True)
    last_factuurnummer = models.PositiveIntegerField(default=0)
    contributie_iban = models.CharField(max_length=34, null=True, blank=True)
    tenaamstelling = models.CharField(max_length=100, null=True, blank=True)
    
    class Meta:
        verbose_name = _("configuration")
        verbose_name_plural = _("configurations")

config = SimpleLazyObject(Configuration.objects.get)

class Activity(models.Model):
    description = models.CharField(max_length=15, unique=True)
    
    class Meta:
        verbose_name = _("activity")
        verbose_name_plural = _("activities")
        ordering = ['description']

    def __str__(self):
        return self.description
    
class Leeftijdscategory(models.Model):
    category = models.CharField(max_length=3, unique=True)
    description = models.CharField(max_length=25)
    maximum_age = models.PositiveIntegerField(_("maximum age end of season"), default=100)
    
    class Meta:
        verbose_name = _("age category")
        verbose_name_plural = _("age categories")
        ordering = ['-maximum_age']

    def __str__(self):
        return self.description
        
class Longcategory(models.Model):
    longcategory = models.CharField(max_length=25, unique=True)
    leeftijdscategory = models.ForeignKey(Leeftijdscategory, on_delete=models.PROTECT, null=True)
    
    class Meta:
        verbose_name = _("extended category name")
        verbose_name_plural = _("extended category names")
        ordering = ['leeftijdscategory']

    def __str__(self):
        return self.longcategory

class PaymentType(models.Model):
    type = models.CharField(max_length=25, unique=True)
    
    class Meta:
        verbose_name = _("payment type")
        verbose_name_plural = _("payment types")
        ordering = ['type']

    def __str__(self):
        return self.type
        
class Paymentmethod(models.Model):
    description = models.CharField(max_length=25, unique=True)
    type = models.ForeignKey(PaymentType, on_delete=models.SET_NULL, null=True)
    
    class Meta:
        verbose_name = _("payment method")
        verbose_name_plural = _("payment methods")
        ordering = ['description']

    def __str__(self):
        return self.description
        
class MemberImport(models.Model):
    """
    Imports of memberfiles
    """
    import_datetime = models.DateTimeField(_("timestamp"), unique=True, auto_now_add=True)
    active_members = models.PositiveIntegerField(_("active members"), default=0)
    new_members = models.PositiveIntegerField(_("new registrations"), default=0)
    unsubscribed_members = models.PositiveIntegerField(_("unsubscribes"), default=0)
    changed_members = models.PositiveIntegerField(_("changes"), default=0)
    non_active_members = models.PositiveIntegerField(_("non active members"), default=0)
    
    class Meta:
        verbose_name = _("member import")
        verbose_name_plural = _("member imports")
        ordering = ['-import_datetime']

    def __str__(self):
        return self.import_datetime.strftime('%Y-%m-%d %H:%M')
        
class Inschrijving(models.Model):
    datumtijd = models.DateTimeField(null=True)
    roepnaam = models.CharField(max_length=50)
    achternaam = models.CharField(max_length=50)
    adres = models.CharField(max_length=50)
    woonplaats = models.CharField(max_length=30)
    postcode = models.CharField(max_length=7)
    geboortedatum = models.DateField()
    machtiging = models.BooleanField(default=False)
    naam_machtiging = models.CharField(max_length=50)
    adres_machtiging = models.CharField(max_length=50)
    postcode_machtiging = models.CharField(max_length=7)
    plaats_machtiging = models.CharField(max_length=30)
    iban = models.CharField(max_length=34, null=True, blank=True)
    privacy_policy = models.BooleanField(default=False)
    webform = models.BooleanField(default=False)
    
    class Meta:
        indexes = [
            models.Index(fields=['geboortedatum', 'postcode']),
        ]
        verbose_name = _("subscription")
        verbose_name_plural = _("subscriptions")

    def __str__(self):
        return ("%s %s" % (self.roepnaam, self.achternaam))
                                
class Memberstatus(models.Model):
    status = models.CharField(max_length=10, unique=True)
    description = models.CharField(max_length=50)
    
    class Meta:
        verbose_name = _("member status")
        verbose_name_plural = _("member statuses")
    
    def __str__(self):
        return self.status

class Member(models.Model):
    relatiecode = models.CharField(max_length=7, unique=True)
    voorletters = models.CharField(max_length=10, blank=True)
    roepnaam = models.CharField(max_length=25, blank=True)
    tussenvoegsels = models.CharField(max_length=10, blank=True)
    achternaam = models.CharField(max_length=50, blank=True)
    geslacht = models.CharField(max_length=1,default='M')
    geboortedatum = models.DateField(blank=True, null=True)
    straatnaam = models.CharField(max_length=50, blank=True)
    huisnummer = models.IntegerField(blank=True, null=True)
    toevoeging = models.CharField(max_length=10, blank=True)
    postcode = models.CharField(max_length=7, blank=True)
    plaats = models.CharField(max_length=30, blank=True)
    telefoon = models.CharField(max_length=15, blank=True)
    mobiel = models.CharField(max_length=15, blank=True)
    email = models.EmailField(blank=True)
    aanmeldingsdatum = models.DateField(blank=True, null=True)
    inschrijvingsdatum = models.DateField(blank=True, null=True)
    machtigingsdatum = models.DateField(blank=True, null=True)
    activities = models.ManyToManyField(Activity)
    inschrijving = models.OneToOneField(Inschrijving, on_delete=models.SET_NULL, null=True)
    naam_machtiging = models.CharField(max_length=50, blank=True)
    adres_machtiging = models.CharField(max_length=50, blank=True)
    postcode_machtiging = models.CharField(max_length=7, blank=True)
    plaats_machtiging = models.CharField(max_length=30, blank=True)
    email_machtiging = models.EmailField(blank=True, null=True)
    kenmerk_machtiging = models.CharField(max_length=50, blank=True)
    # korting percentage blijft behouden, ook in nieuwe seizoen
    # termijnbetalingen moet ieder seizoen opnieuw afgesproken worden, dus niet in deze tabel
    kortingpercentage = models.PositiveIntegerField(default=0)
    huygenspas = models.BooleanField(default=False)
    machtiging = models.BooleanField(default=False)
    machtiging_withdrawn = models.BooleanField(default=False)
    payment_method = models.ForeignKey(Paymentmethod, on_delete=models.PROTECT, null=True)
    iban = models.CharField(max_length=34, null=True, blank=True)
    last_import = models.ForeignKey(MemberImport, on_delete=models.SET_NULL, blank=True, null=True)
    status = models.ForeignKey(Memberstatus, on_delete=models.SET_NULL, blank=True, null=True)
    # calculated fields
    lc = models.ForeignKey(Leeftijdscategory, on_delete=models.SET_NULL, blank=True, null=True)

    @ property
    def formalname(self):
        if self.tussenvoegsels:
            return '%s %s %s' % ((self.voorletters or self.roepnaam[0]), self.tussenvoegsels, self.achternaam)
        else:
            return '%s %s' % ((self.voorletters or self.roepnaam[0]), self.achternaam)

    @ property
    def fullname(self):
        if self.tussenvoegsels:
            return '%s %s %s' % ((self.roepnaam or self.voorletters), self.tussenvoegsels, self.achternaam)
        else:
            return '%s %s' % ((self.roepnaam or self.voorletters), self.achternaam)

    @ property
    def fulladdress(self):
        return '%s %s%s, %s %s' % (self.straatnaam, str(self.huisnummer), self.toevoeging, self.postcode, self.plaats)

    @ property
    def shortaddress(self):
        return '%s %s%s' % (self.straatnaam, str(self.huisnummer), self.toevoeging)

    @ property
    def leeftijdscategorie(self):
        if not self.geboortedatum:
            return Leeftijdscategory.objects.get(category='Sen')
        startleeftijd = config.seizoen.startjaar - self.geboortedatum.year
        if startleeftijd > 18:
            categorie = Leeftijdscategory.objects.get(category='Sen')
        elif startleeftijd > 16:
            categorie = Leeftijdscategory.objects.get(category='O19')
        elif startleeftijd > 14:
            categorie = Leeftijdscategory.objects.get(category='O17')
        elif startleeftijd > 12:
            categorie = Leeftijdscategory.objects.get(category='O15')
        elif startleeftijd > 10:
            categorie = Leeftijdscategory.objects.get(category='O13')
        elif startleeftijd > 8:
            categorie = Leeftijdscategory.objects.get(category='O11')
        elif startleeftijd > 6:
            categorie = Leeftijdscategory.objects.get(category='O9')
        else:
            categorie = Leeftijdscategory.objects.get(category='O7')
        return categorie

    def matchInschrijving(self, i):
        self.inschrijving = i
        self.naam_machtiging = i.naam_machtiging
        self.adres_machtiging = i.adres_machtiging
        self.postcode_machtiging = i.postcode_machtiging
        self.plaats_machtiging = i.plaats_machtiging
        self.machtiging = i.machtiging
        self.iban = i.iban
        if valid_iban(self.iban):
            self.machtiging = True
        if self.huygenspas:
            self.payment_method = Paymentmethod.objects.get(description='Huygenspas')
        else:
            if valid_iban(self.iban) and self.machtiging:
                self.payment_method = Paymentmethod.objects.get(description='Incasso')
            else:
                self.payment_method = Paymentmethod.objects.get(description='Factuur')
        self.save()
        for c in self.contributions.all():
            c.factuur_naam = self.naam_machtiging
            c.factuur_adres = self.adres_machtiging
            c.factuur_postcode = self.postcode_machtiging
            c.factuur_plaats = self.plaats_machtiging
            c.iban = self.iban
            c.payment_method = self.payment_method
            c.save()
            c.recreate_payments()

    def recalc(self):
        self.lc = self.leeftijdscategorie

    def get_absolute_url(self):
        return reverse('jarrbo_contributie:member_detail', kwargs={'pk': self.pk})
    
    class Meta:
        verbose_name = _("member")
        verbose_name_plural = _("members")
        ordering = ['achternaam', 'tussenvoegsels', 'roepnaam']
        indexes = [models.Index(fields=['relatiecode']),
                   models.Index(fields=['postcode']),
                   models.Index(fields=['geboortedatum']),
                   models.Index(fields=['last_import']),
                   ]

    def save(self, *args, **kwargs):
        self.recalc()
        super(Member, self).save(*args, **kwargs)

    def __str__(self):
        return self.fullname

class ContributionManager(models.Manager):
    def get_queryset(self):
        return super().get_queryset().filter(seizoen__current=True)

    def create_contribution(self, member, seizoen, activity):
        c = self.create(member=member, seizoen=seizoen, activity=activity)
        if activity.description == 'Donateur':
            method = Paymentmethod.objects.get(description='Factuur')
            c.payment_method = method
            c.recreate_payments()
        if activity.description != 'Donateur':
            same_adres = Member.objects.filter(postcode=member.postcode,huisnummer=member.huisnummer).count()
            adres_korting = {2: 15, 3: 20, 4:25}
            c.kortingopadres = adres_korting.get(same_adres, 0)
        c.save()
        return c

class Contribution(models.Model):
    member = models.ForeignKey(Member, on_delete=models.PROTECT, related_name='contributions')
    seizoen = models.ForeignKey(Seizoen, on_delete=models.PROTECT, null=True)
    activity = models.ForeignKey(Activity, on_delete=models.PROTECT, related_name='contributions')
    kortingopadres = models.DecimalField(max_digits=6, decimal_places=2, default=0)
    kortingpercentage = models.PositiveIntegerField(default=0)
    kortingvast = models.DecimalField(max_digits=6, decimal_places=2, default=0)
    aanmaningskosten = models.DecimalField(max_digits=6, decimal_places=2, default=0)
    termijnen = models.PositiveIntegerField(default=1)
    payment_method = models.ForeignKey(Paymentmethod, on_delete=models.PROTECT, null=True)
    iban = models.CharField(max_length=34, null=True, blank=True)
    factuur_naam = models.CharField(max_length=50, null=True, blank=True)
    factuur_adres = models.CharField(max_length=50, null=True, blank=True)
    factuur_postcode = models.CharField(max_length=7, null=True, blank=True)
    factuur_plaats = models.CharField(max_length=30, null=True, blank=True)
    factuur_email = models.EmailField(null=True, blank=True)
    sponsored = models.BooleanField(default=False)
    # calculated fields
    # korting percentage op aanmelddatum
    kdp = models.PositiveIntegerField(default=0)
    # basis contributie
    bc = models.DecimalField(max_digits=6, decimal_places=2, default=0)
    # kledingfonds
    kf = models.DecimalField(max_digits=6, decimal_places=2, default=0)
    # administratiekosten
    ak = models.DecimalField(max_digits=6, decimal_places=2, default=0)
    # kortingopdatum
    kd = models.DecimalField(max_digits=6, decimal_places=2, default=0)
    # korting speciaal
    ks = models.DecimalField(max_digits=6, decimal_places=2, default=0)
    # totaal contributiebedrag
    tc = models.DecimalField(max_digits=6, decimal_places=2, default=0)
    # totaal ontvangen
    received = models.DecimalField(max_digits=6, decimal_places=2, default=0)

    @ property
    def kortingdatum(self):
        if not self.member.aanmeldingsdatum:
            return 0
        startjaar = config.seizoen.startjaar
        eindjaar = config.seizoen.eindjaar
        if self.member.aanmeldingsdatum.year == startjaar and self.member.aanmeldingsdatum.month >= 11:
            return 20
        if self.member.aanmeldingsdatum.year == eindjaar:
            if self.member.aanmeldingsdatum.month < 3:
                return 50
            else:
                if self.member.aanmeldingsdatum.month < 4:
                    return 60
                else:
                    return 70
        return 0

    @ property
    def base_contribution(self):
        if self.sponsored:
            return decimal.Decimal(0.00)
        ct = ContributionTable.seizoen_objects.get(activity=self.activity,
                                       leeftijdscategorie=self.member.leeftijdscategorie)
        return ct.base_contribution

    @ property
    def kledingfonds(self):
        if self.sponsored:
            return decimal.Decimal(0.00)
        ct = ContributionTable.seizoen_objects.get(activity=self.activity,
                                       leeftijdscategorie=self.member.leeftijdscategorie)
        return ct.kledingfonds

    @ property
    def administratiekosten(self):
        if self.activity.description == 'Donateur':
            return decimal.Decimal(0.00)
        if self.sponsored:
            return decimal.Decimal(0.00)
        if not self.payment_method or \
               self.payment_method.description == 'Incasso' or \
               self.payment_method.description == 'Jeugdsportfonds':
            return decimal.Decimal(0.00)
        else:
            return decimal.Decimal(7.50)

    @ property
    def korting(self):
        if self.sponsored:
            return decimal.Decimal(0.00)
        if self.kortingpercentage > 0:
            return round((self.base_contribution * self.kortingpercentage) / 100, 2)
        else:
            return self.kortingvast

    @ property
    def kortingopdatum(self):
        if self.sponsored:
            return decimal.Decimal(0.00)
        if self.kortingdatum > 0:
            return round(((self.base_contribution - self.korting) * self.kortingdatum) / 100, 2)
        else:
            return decimal.Decimal(0.00)
    
    @ property
    def total_cost(self):
        if self.sponsored:
            return decimal.Decimal(0.00)
        c = self.aanmaningskosten \
            + self.administratiekosten
        return c
    
    @ property
    def total_korting(self):
        if self.sponsored:
            return decimal.Decimal(0.00)
        c = self.kortingopadres \
            + self.kortingopdatum \
            + self.korting
        return c
    
    @ property
    def total_contribution(self):
        if self.sponsored:
            return decimal.Decimal(0.00)
        c = self.base_contribution \
            + self.kledingfonds \
            + self.total_cost \
            - self.total_korting
        return c

    @ property
    def payed(self):
        payments = Payment.seizoen_objects.filter(contribution=self,status__status='Betaald')
        p = payments.aggregate(total=Sum('amount'))['total']
        return p or decimal.Decimal(0.00)

    @ property
    def sent(self):
        payments = Payment.seizoen_objects.filter(contribution=self,status__status='Verzonden')
        p = payments.aggregate(total=Sum('amount'))['total']
        return p or decimal.Decimal(0.00)

    @ property
    def planned(self):
        payments = Payment.seizoen_objects.filter(contribution=self,status__status='Gepland')
        p = payments.aggregate(total=Sum('amount'))['total']
        return p or decimal.Decimal(0.00)

    def recreate_payments(self, fromdate=None):
        # Do not create payments when status of member is aangemeld (not playing yet)
        if not self.member.status or self.member.status.description == 'Aangemeld':
            return
        # first remove all open payments of current contribution
        payments = self.payments.filter(status__status='Gepland')
        if fromdate:
            payments = payments.filter(paymentdate__gt=fromdate)
        payments.filter(status__status='Gepland').delete()
        if self.sponsored:
            if self.payed + self.sent > 0:
                p = Payment(seizoen = self.seizoen,
                            contribution = self,
                            method = Paymentmethod.objects.get(description='Overboeking'),
                            amount = - (self.payed + self.sent),
                            status = Paymentstatus.objects.get(status='Gepland'),
                            createdate = datetime.today(),
                            )
                p.save()
            return
        remaining = self.total_contribution - self.payed - self.sent - self.planned
        if not remaining:
            return
        termijnen_betaald = self.payments.filter(status__include=True).count()
        termijnen_planned = self.payments.filter(status__status='Gepland').count()
        termijnen_left = self.termijnen - termijnen_betaald - termijnen_planned
        if remaining and termijnen_left == 0:
            termijnen_left = 1
        st = PaymentbatchStatus.objects.get(status='Gepland')
        batches = Paymentbatch.seizoen_objects.filter(status=st)
        if fromdate:
            batches = batches.filter(datum__gt=fromdate)
        batches = batches.order_by('datum')
        if termijnen_left > batches.count():
            termijnen_left = batches.count()
        if termijnen_left < 1:
            i = 1
        else:
            i = termijnen_left
        if len(batches) > 0:
            for batch in batches[:i]:
                amount = round(remaining / i, 2)
                remaining = remaining - amount
                if self.payment_method.description == 'Incasso':
                    paymentbatch = batch
                    paymentdate = batch.datum
                else:
                    paymentbatch = None
                    paymentdate = batch.datum
                p = Payment(seizoen = self.seizoen,
                            contribution = self,
                            paymentbatch = paymentbatch,
                            method = self.payment_method,
                            amount = amount,
                            paymentdate = paymentdate,
                            status = Paymentstatus.objects.get(status='Gepland'),
                            createdate = datetime.today(),
                            )
                i = i - 1
                p.save()
                

    def update_memberdata(self):
        self.factuur_naam = self.member.naam_machtiging
        self.factuur_adres = self.member.adres_machtiging
        self.factuur_postcode = self.member.postcode_machtiging
        self.factuur_plaats = self.member.plaats_machtiging
        self.factuur_email = self.member.email_machtiging
        self.iban = self.member.iban
        self.kortingpercentage = self.member.kortingpercentage
        self.payment_method = self.member.payment_method
        self.save()
        self.recreate_payments()

    def recalc(self):
        # recalculate dependent fields
        self.kdp = self.kortingdatum
        self.bc = self.base_contribution
        self.kf = self.kledingfonds
        self.ak = self.administratiekosten
        if self.kortingpercentage > 0:
            self.ks = round((self.bc * self.kortingpercentage) / 100, 2)
        else:
            self.ks = self.kortingvast
        if self.kdp > 0:
            self.kd = round(((self.bc - self.ks) * self.kdp) / 100, 2)
        else:
            self.kd = decimal.Decimal(0.00)
        if self.sponsored:
            self.tc = decimal.Decimal(0.00)
        else:
            self.tc = self.bc + self.aanmaningskosten + self.ak + self.kf \
                      - self.kortingopadres - self.kd - self.ks
        payments = Payment.seizoen_objects.filter(contribution=self,status__status='Betaald')
        p = payments.aggregate(total=Sum('amount'))['total']
        self.received = p or decimal.Decimal(0.00)

    
    class Meta:
        verbose_name = _("contribution")
        verbose_name_plural = _("contributions")
        unique_together = [['member', 'seizoen', 'activity']]
        ordering = ['member', 'activity']
        
    seizoen_objects = ContributionManager()
    objects = models.Manager()

    def get_absolute_url(self):
        return reverse('jarrbo_contributie:contribution_detail', kwargs={'pk': self.pk})

    def save(self, *args, **kwargs):
        self.recalc()
        super(Contribution, self).save(*args, **kwargs)

    def __str__(self):
        return ("%s, %s" % (self.member.fullname, self.activity.description))

class Machtiging(models.Model):        
    datumtijd = models.DateTimeField(unique=True)
    voornaam = models.CharField(max_length=50)
    achternaam = models.CharField(max_length=50)
    adres = models.CharField(max_length=50)
    woonplaats = models.CharField(max_length=30)
    postcode = models.CharField(max_length=7)
    geboortedatum = models.DateField()
    naam_machtiging = models.CharField(max_length=50)
    plaats_machtiging = models.CharField(max_length=30)
    email_machtiging = models.EmailField(blank=True, null=True)
    iban = models.CharField(max_length=34, blank=True)
    
    class Meta:
        indexes = [
            models.Index(fields=['geboortedatum', 'postcode']),
        ]
        verbose_name = _("authority")
        verbose_name_plural = _("authorities")

    def __str__(self):
        return ("%s %s" % (self.voornaam, self.achternaam))

class ContributionTableManager(models.Manager):
    def get_queryset(self):
        return super().get_queryset().filter(seizoen__current=True)

class ContributionTable(models.Model):
    seizoen = models.ForeignKey(Seizoen, on_delete=models.PROTECT)
    activity = models.ForeignKey(Activity, on_delete=models.PROTECT)
    leeftijdscategorie = models.ForeignKey(Leeftijdscategory, on_delete=models.PROTECT)
    base_contribution = models.DecimalField(max_digits=6, decimal_places=2, default=0)
    kledingfonds = models.DecimalField(max_digits=6, decimal_places=2, default=0)
    
    class Meta:
        verbose_name = _("contribution amount")
        verbose_name_plural = _("contribution amounts")
        ordering = ['seizoen', 'activity', 'leeftijdscategorie']
        unique_together = [['seizoen', 'activity', 'leeftijdscategorie']]

    seizoen_objects = ContributionTableManager()
    objects = models.Manager()

    def __str__(self):
        return ("%s: %s %s" % (self.seizoen.description, self.leeftijdscategorie. description, \
                                self.activity.description))
                               
class Paymentstatus(models.Model):
    status = models.CharField(max_length=15, unique=True)
    include = models.BooleanField(default=False)
    description = models.CharField(max_length=25)
    regular = models.BooleanField(default=True,
                help_text="part of standard proces, if not new payment must be created")
    actie = models.CharField(max_length=25, null=True)
    
    class Meta:
        verbose_name = _("payment status")
        verbose_name_plural = _("payment statuses")

    def __str__(self):
        return self.description

class PaymentStatusCode(models.Model):
    paymentstatus = models.ForeignKey(Paymentstatus, null=True, on_delete=models.PROTECT, 
                                        related_name='codes')
    status_code = models.CharField(max_length=4, unique=True)
    description = models.CharField(max_length=60)
    long_description = models.TextField(null=True, blank=True)
    
    class Meta:
        verbose_name = _("payment status code")
        verbose_name_plural = _("payment status codes")
        ordering = ['status_code', ]

    def __str__(self):
        return ("%s-%s-%s" % (self.paymentstatus, self.status_code, self.description,))

class PaymentbatchStatus(models.Model):
    status = models.CharField(max_length=15, unique=True)
    
    class Meta:
        verbose_name = _("paymentbatch status")
        verbose_name_plural = _("paymentbatch statuses")

    def __str__(self):
        return self.status

class PaymentbatchManager(models.Manager):
    def get_queryset(self):
        return super().get_queryset().filter(seizoen__current=True)
        
    def nextbatch(self):
        planned = self.filter(status__status='Gepland')
        if planned:
            return planned[0]
        else:
            return None

class Paymentbatch(models.Model):
    seizoen = models.ForeignKey(Seizoen, on_delete=models.PROTECT)
    datum = models.DateField()
    status = models.ForeignKey(PaymentbatchStatus, null=True, on_delete=models.SET_NULL)

    @ property
    def withdrawn(self):
        return self.payments.filter(status__status='Teruggeboekt').count()

    @ property
    def nobalance(self):
        return self.payments.filter(status__status='Geensaldo').count()

    @ property
    def totalplanned(self):
        return self.payments.filter(status__status='Gepland').aggregate(totalamount=Sum('amount'))['totalamount']

    @ property
    def totalpayed(self):
        return self.payments.filter(status__status='Betaald').aggregate(totalamount=Sum('amount'))['totalamount']

    @ property
    def totalwithdrawn(self):
        return self.payments.filter(status__status='Teruggeboekt').aggregate(totalamount=Sum('amount'))['totalamount']

    @ property
    def totalnosaldo(self):
        return self.payments.filter(status__status='Geensaldo').aggregate(totalamount=Sum('amount'))['totalamount']
        
    def createIncassobatch(self):
        from openpyxl import load_workbook
        from django.conf import settings
        from unidecode import unidecode
        path = '/'.join((settings.MEDIA_ROOT,'jarrbo_contributie','Incassobatch_template_v1.1.xlsx'))
        wb = load_workbook(path)
        ws = wb.active
        ws['C4'] = self.datum
        row = 12
        klfnds = ' (dit is incl. bijdrage kledingfonds)'
        # to be sure include ony the payments that has Incasso as method
        for payment in self.payments.filter(method__description='Incasso'):
            tenaamstellingref = ws.cell(row=row, column=2)
            if payment.contribution.factuur_naam:
                tenaamstellingref.value = unidecode(payment.contribution.factuur_naam)
            else:
                tenaamstellingref.value = unidecode(payment.contribution.member.formalname)
            ibanref = ws.cell(row=row, column=3)
            ibanref.value = payment.contribution.iban
            kenmerkref = ws.cell(row=row, column=4)
#            16-08-2022 change kenmerk to solve issues with ongeldige banktransactie code
#            kenmerkref.value = payment.contribution.member.kenmerk_machtiging or \
#                                payment.contribution.member.relatiecode
            kenmerkref.value = ("%s-%s" % (payment.contribution.member.relatiecode,
                                      payment.contribution.activity.description[0:1]))
            bedragref = ws.cell(row=row, column=5)
            bedragref.value = payment.amount
            omschrijvingref = ws.cell(row=row, column=6)
            omschrijvingref.value = unidecode(payment.remark)
            machtigingdatumref = ws.cell(row=row, column=7)
            machtigingdatumref.value = payment.contribution.member.machtigingsdatum or \
                                        payment.contribution.member.inschrijvingsdatum or \
                                        payment.contribution.member.aanmeldingsdatum
            row += 1
        return wb
    
    class Meta:
        verbose_name = _("payment batch")
        verbose_name_plural = _("payment batches")
        ordering = ['seizoen', 'datum']

    seizoen_objects = PaymentbatchManager()
    objects = models.Manager()

    def __str__(self):
        return ("%s" % (self.datum.strftime('%d %b'),))

class PaymentManager(models.Manager):
    def get_queryset(self):
        return super().get_queryset().filter(seizoen__current=True)

class Payment(models.Model):
    seizoen = models.ForeignKey(Seizoen, on_delete=models.PROTECT)
    contribution = models.ForeignKey(Contribution, on_delete=models.PROTECT, related_name='payments')
    paymentbatch = models.ForeignKey(Paymentbatch, blank=True, null=True, on_delete=models.PROTECT, related_name='payments')
    method = models.ForeignKey(Paymentmethod,on_delete=models.PROTECT)
    amount = models.DecimalField(max_digits=6, decimal_places=2, default=0)
    status = models.ForeignKey(Paymentstatus,on_delete=models.PROTECT)
    aanvraagnummer = models.CharField(max_length=10, blank=True, null=True)
    factuurnummer = models.PositiveIntegerField(null=True)
    createdate = models.DateField(blank=True, null=True)
    submitdate = models.DateField(blank=True, null=True)
    paymentdate = models.DateField(blank=True, null=True)
    withdrawndate = models.DateField(blank=True, null=True)
    withdrawnmaildate = models.DateField(blank=True, null=True)
    huygensmaildate = models.DateField(blank=True, null=True)
    duedate = models.DateField(blank=True, null=True)
    paymentstatuscode = models.ForeignKey(PaymentStatusCode, blank=True, null = True, on_delete=models.SET_NULL)

    seizoen_objects = PaymentManager()
    objects = models.Manager()

    @ property
    def remark(self):
        act = ''
        incl = ''
        klfnds = ''
        kort = ''
        kost = ''
        en = ''
        if self.contribution.activity.description == 'Zaal':
            act = ' (Zaal)'
        if self.contribution.kledingfonds > 0:
            incl = ', incl'
            klfnds = ' kledingfonds'
        if self.contribution.korting > 0 or \
           self.contribution.kortingvast > 0 or \
           self.contribution.kortingopdatum > 0 or \
           self.contribution.kortingvast > 0:
            if klfnds:
                en = ' en'
            incl = ', incl'
            kort = ' korting'
        if self.contribution.administratiekosten > 0 or \
           self.contribution.aanmaningskosten > 0:
            if klfnds:
                en = ' en'
            incl = ', incl'
            if kort:
                kost = ' en kosten'
            else:
                kost = ' kosten'
        return ("%s%s%s%s%s%s%s" % (self.contribution.member.roepnaam, act,
                                    incl, klfnds, en, kort, kost))

    def get_absolute_url(self):
        return reverse('jarrbo_contributie:payment_detail', kwargs={'pk': self.pk})
    
    class Meta:
        verbose_name = _("payment")
        verbose_name_plural = _("payments")
        ordering = ['seizoen',
                    'contribution__member__achternaam',
                    'contribution__member__tussenvoegsels',
                    'contribution__member__roepnaam',
                    'paymentbatch__datum',
                    'paymentdate',
                   ]

    def __str__(self):
        return ("%s %s" % ("{0:.2f}".format(self.amount), self.status.status))
        
class PaymentstatusChange(models.Model):
    paymenttype = models.ForeignKey(PaymentType, null=True, on_delete=models.SET_NULL)
    statusbefore = models.ForeignKey(Paymentstatus, null=True, on_delete=models.SET_NULL,
                                    related_name='statusbefore')
    statusafter = models.ForeignKey(Paymentstatus, null=True, on_delete=models.SET_NULL,
                                    related_name='statusafter')
    standard = models.BooleanField(default=True)
    recovery = models.BooleanField(default=False)

    class Meta:
        unique_together = [['paymenttype', 'statusbefore', 'statusafter']]
        verbose_name = _("payment status change")
        verbose_name_plural = _("payment status changes")

    def __str__(self):
        return ("%s-%s-%s" % (self.paymenttype, self.statusbefore, self.statusafter))
        
class Note(models.Model):
    member = models.ForeignKey(Member, null=True, on_delete=models.SET_NULL, related_name='notes')
    datetime = models.DateTimeField(_("timestamp"), unique=True, auto_now_add=True)
    description = models.CharField(max_length=150, null=True)
    note = models.TextField(null=True, blank=True)

    class Meta:
        verbose_name = _("notes")
        verbose_name_plural = _("notes")
        ordering = ['-datetime', ]

    def __str__(self):
        return self.description

class CoronaRestitution(models.Model):
    contribution = models.OneToOneField(Contribution, null=True, on_delete=models.SET_NULL)
    s_2021 = models.BooleanField(default=False)
    applied = models.BooleanField(default=False)
    amount = models.DecimalField(max_digits=6, decimal_places=2, default=0)
    payed = models.BooleanField(default=False)
    
    class Meta:
        verbose_name = _("payment")
        verbose_name_plural = _("payments")
        ordering = [
                    'contribution__member__achternaam',
                    'contribution__member__tussenvoegsels',
                    'contribution__member__roepnaam',
                   ]
